"""Extrae el contenido del Excel 2026_Budget_Template.xlsx al seed JSON de la app.

Uso: python3 scripts/extract_seed.py <ruta_xlsx> <ruta_salida_json>
"""
import json
import re
import sys
import unicodedata

import openpyxl

MONTHS = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
          'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
YEAR = 2026
# Bloques del registro semanal: (concepto, categoria, aed, eur, fecha) por semana
WEEK_BLOCKS = [(1, 2, 3, 4, 5), (6, 7, 8, 9, 10), (11, 12, 13, 14, 15),
               (16, 17, 18, 19, 20), (21, 22, 23, 24, 0)]


def slug(text):
    text = unicodedata.normalize('NFKD', str(text)).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')


def num(value):
    return round(float(value), 2) if isinstance(value, (int, float)) else None


def main(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path)
    resumen, ahorros = wb['Resumen'], wb['Ahorros']

    rate = resumen['D1'].value
    categories = [c.value.strip() for c in wb['_Categorías']['A'] if c.value]
    # El Excel escribe categorías con espacios sobrantes o en singular
    alias = {slug(c): c for c in categories}
    alias['restaurante'] = 'Restaurantes'

    def category_of(raw):
        if not raw:
            return 'Otros'
        return alias.get(slug(raw), str(raw).strip())

    # Los nombres reales (Salario Santi/Bea) están en las hojas mensuales, no en Resumen
    labels = [wb['Enero'].cell(r, 1).value for r in range(3, 6)]
    incomes = [{'id': slug(resumen.cell(r, 1).value), 'name': label,
                'amount': num(resumen.cell(r, 3).value), 'currency': 'AED'}
               for r, label in zip(range(4, 7), labels)]
    fixed = [{'id': slug(resumen.cell(r, 1).value), 'name': resumen.cell(r, 1).value,
              'amount': num(resumen.cell(r, 2).value), 'currency': 'AED'} for r in range(11, 26)]

    def savings_rows(rows):
        return [{'id': slug(ahorros.cell(r, 1).value), 'name': ahorros.cell(r, 1).value,
                 'goal': num(ahorros.cell(r, 2).value) or 0,
                 'contribution': num(ahorros.cell(r, 3).value) or 0,
                 'initialBalance': num(ahorros.cell(r, 4).value) or 0} for r in rows]

    transactions, monthIncomes = [], {}
    for index, name in enumerate(MONTHS, start=1):
        ws = wb[name]
        month = f'{YEAR}-{index:02d}'

        # Ingresos del mes: el Excel sustituye la fórmula por un número cuando se ajusta a mano
        overrides = []
        for row, base in zip(range(3, 6), incomes):
            aed, eur = ws.cell(row, 3).value, ws.cell(row, 4).value
            if isinstance(eur, (int, float)):
                overrides.append({**base, 'name': ws.cell(row, 1).value, 'amount': num(eur), 'currency': 'EUR'})
            elif isinstance(aed, (int, float)):
                overrides.append({**base, 'name': ws.cell(row, 1).value, 'amount': num(aed), 'currency': 'AED'})
        if overrides:
            monthIncomes[month] = overrides

        header = next(r for r in range(33, ws.max_row + 1) if ws.cell(r, 1).value == 'Concepto')
        end = next(r for r in range(header + 1, ws.max_row + 1) if ws.cell(r, 1).value == 'TOTAL')
        for week, (c_con, c_cat, c_aed, c_eur, c_date) in enumerate(WEEK_BLOCKS, start=1):
            for row in range(header + 1, end):
                concept = ws.cell(row, c_con).value
                aed, eur = num(ws.cell(row, c_aed).value), num(ws.cell(row, c_eur).value)
                if not concept and aed is None and eur is None:
                    continue
                date = ws.cell(row, c_date).value if c_date else None
                transactions.append({
                    'id': f'{month}-w{week}-r{row}',
                    'month': month,
                    'week': week,
                    'date': date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else None,
                    'concept': (str(concept).strip() if concept else 'Sin concepto'),
                    'category': category_of(ws.cell(row, c_cat).value),
                    'amount': aed if aed is not None else (eur or 0),
                    'currency': 'AED' if aed is not None else 'EUR',
                })

    for tx in transactions:
        if tx['category'] not in categories:
            categories.append(tx['category'])

    seed = {
        'version': 1,
        'year': YEAR,
        'settings': {'rate': rate, 'categories': categories, 'incomes': incomes, 'fixedExpenses': fixed},
        'savings': {
            'accounts': savings_rows(range(5, 10)),
            'emergency': savings_rows(range(11, 14)),
            'sinking': savings_rows(range(18, 19)),
        },
        'monthIncomes': monthIncomes,
        'transactions': transactions,
    }
    with open(out_path, 'w', encoding='utf-8') as fh:
        json.dump(seed, fh, ensure_ascii=False, indent=2)

    print(f'{len(transactions)} movimientos, {len(categories)} categorías, tipo {rate}')
    print('ingresos ajustados por mes:', list(monthIncomes))


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
